#!/usr/bin/env python3
"""
Convert MS Project CSV export to XLSX format
"""
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Read the CSV file
csv_file = 'from-pm-tools/ms-project-export.csv'
xlsx_file = 'from-pm-tools/ms-project-export.xlsx'

wb = Workbook()
ws = wb.active
ws.title = "Tasks"

# Read CSV and write to Excel
with open(csv_file, 'r', encoding='utf-8') as f:
    reader = csv.reader(f)
    for row_idx, row in enumerate(reader, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Format header row
            if row_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

# Auto-adjust column widths
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws.column_dimensions[column_letter].width = adjusted_width

# Save the workbook
wb.save(xlsx_file)
print(f"Created {xlsx_file}")
